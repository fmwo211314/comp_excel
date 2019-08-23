# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
# wb = Workbook()
#
# ws1 = wb.create_sheet("Mysheet")           #创建一个sheet

wb = load_workbook("sample.xlsx", data_only=False)
ws1 = wb.active


# #操作单列
# print(ws1["A"])
# for cell in ws1["A"]:
#     print (cell.value)
#
# #操作多列,获取每一个值
# print (ws1["A:C"])
# for column in ws1["A:C"]:
#     for cell in column:
#         print (cell.value)
#
# #操作多行
# row_range = ws1[1:3]
# print( row_range)
# for row in row_range:
#     for cell in row:
#         print( cell.value)

# print ("*"*50)
# for row in ws1.iter_rows(min_row=1, min_col=1, max_col=4, max_row=4):
#     for cell in row:
#         # print(dir(cell))
#         # print(cell.data_type)
#         # print (cell.value)
#         print("{0},{1}".format(cell.data_type,cell.value))

# #获取所有行
# print (ws1.rows)
# for row in ws1.rows:
#     print( row)

# print( "*"*50)
# #获取所有列
# print( ws1.columns)
# for col in ws1.columns:
#     print( col)

# wb.save("sample.xlsx")



ss = "=C11/$C$5"

a = ss.replace("$","")

print(a)



