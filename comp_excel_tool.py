import sys, traceback
from form import Ui_Form
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment


class MyPyQT_Form(QtWidgets.QWidget, Ui_Form):
    def __init__(self):
        super(MyPyQT_Form, self).__init__()
        self.setupUi(self)
        self.excel_expect = ""
        self.excel_actual = ""

    def execute_compare(self):
        if self.excel_actual == "" or self.excel_expect == "":
            self.alert("请选择要比较的Excel文件")
            return
        try:
            if self.comboBox.currentText() == "设置范围":
                row_from = int(self.lineEdit_4.text())
                row_to = int(self.lineEdit_3.text())
                column_from = int(self.lineEdit_6.text())
                column_to = int(self.lineEdit_5.text())
                list = [row_from, row_to, column_from, column_to]
                if all(list):
                    self.compareExecuteFromTo(self.excel_expect, self.excel_actual, row_from, row_to, column_from,
                                              column_to)
                else:
                    self.alert("请设置比较范围")
                    return
            else:
                self.compareExecute(self.excel_expect, self.excel_actual)

            self.textBrowser.append("比较完成，详情请查看\n%s" % self.excel_actual)
            self.alert("执行完成")
        except:
            msg = traceback.format_exc()
            self.textBrowser.setText(msg)

    def alert(self, str):
        msg_box = QtWidgets.QMessageBox
        msg_box.information(self, '提示', str, msg_box.Ok)

    def select_expect(self):
        openfile_name = QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel files(*.xlsx , *.xls)')
        self.lineEdit.setText(openfile_name[0])
        self.excel_expect = openfile_name[0]

    def select_actual(self):
        openfile_name = QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel files(*.xlsx , *.xls)')
        self.lineEdit_2.setText(openfile_name[0])
        self.excel_actual = openfile_name[0]

    def select_method(self, index):
        if self.comboBox.currentText() == "不设置":
            self.frame.hide()
        else:
            self.frame.show()

    def select_clear(self):
        self.excel_expect = ""
        self.excel_actual = ""
        self.textBrowser.clear()
        self.lineEdit.clear()
        self.lineEdit_2.clear()
        self.lineEdit_3.clear()
        self.lineEdit_4.clear()
        self.lineEdit_5.clear()
        self.lineEdit_6.clear()

    def compareExecute(self, excel_expect, excel_actual):
        wb_expect = load_workbook(excel_expect, data_only=False)
        wb_actual = load_workbook(excel_actual, data_only=False)
        sheet_expect = wb_expect.active
        sheet_actual = wb_actual.active

        list = []
        for row in sheet_expect.rows:
            for cell in row:
                list.append(cell.value)

        fill = PatternFill("solid", fgColor="FF0000")

        for index_row, row in enumerate(sheet_actual.rows):
            for index_column, cell in enumerate(row):
                index = index_row * sheet_actual.max_column + index_column
                if cell.value != list[index]:
                    try:
                        comment = Comment('预期是：{0},\n实际结果是：{1}'.format(list[index], cell.value), 'lxy')
                        comment.width = 400
                        comment.height = 400
                        cell.comment = comment
                        cell.fill = fill
                        self.textBrowser.append(
                            '第{2}行第{3}列，\n预期是：{0},\n实际结果是：{1}'.format(list[index], cell.value, index_row + 1,
                                                                      index_column + 1))
                    except:
                        pass

        wb_actual.save(excel_actual)

    def compareExecuteFromTo(self, excel_expect, excel_actual, row_from, row_to, column_from, column_to):
        wb_expect = load_workbook(excel_expect, data_only=False)
        wb_actual = load_workbook(excel_actual, data_only=False)
        sheet_expect = wb_expect.active
        sheet_actual = wb_actual.active

        list = []
        for row in sheet_expect.iter_rows(min_row=row_from, min_col=column_from, max_col=column_to,
                                          max_row=row_to):
            for cell in row:
                list.append(cell.value)

        fill = PatternFill("solid", fgColor="FF0000")

        for row in sheet_actual.iter_rows(min_row=row_from, min_col=column_from, max_col=column_to,
                                          max_row=row_to):
            for cell in row:
                index = (cell.row - row_from) * (column_to - column_from + 1) + cell.column - column_from
                result_expect = list[index]
                result_actual = cell.value
                if result_actual != result_expect:
                    if result_expect is not None and result_actual is not None:
                        if result_expect.startswith("="):
                            if "$" in result_expect:
                                result_expect = result_expect.replace("$", "")
                            if result_expect[1:] not in result_actual:
                                try:
                                    comment = Comment('预期是：{0},\n实际结果是：{1}'.format(list[index], cell.value), 'lxy')
                                    comment.width = 400
                                    comment.height = 400
                                    cell.comment = comment
                                    cell.fill = fill
                                    self.textBrowser.append(
                                        '第{2}行第{3}列，\n预期是：{0},\n实际结果是：{1}'.format(list[index], cell.value, cell.row + 1,
                                                                                  cell.column + 1))
                                except:
                                    pass
                    else:
                        try:
                            comment = Comment('预期是：{0},\n实际结果是：{1}'.format(list[index], cell.value), 'lxy')
                            comment.width = 400
                            comment.height = 400
                            cell.comment = comment
                            cell.fill = fill
                            self.textBrowser.append(
                                '第{2}行第{3}列，\n预期是：{0},\n实际结果是：{1}'.format(list[index], cell.value, cell.row + 1,
                                                                          cell.column + 1))
                        except:
                            pass
        wb_actual.save(excel_actual)


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    my_pyqt_form = MyPyQT_Form()
    my_pyqt_form.show()
    sys.exit(app.exec_())
